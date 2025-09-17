#!/usr/bin/env python3
"""
Configuration Manager for Team Task Planner
Handles loading, validating, and managing dynamic configuration for team and project setup.
"""

import json
import os
from typing import Dict, List, Any, Optional
from datetime import datetime

class ConfigManager:
    """Manages configuration for the Team Task Planner."""
    
    def __init__(self, config_path: str = "config.json"):
        """Initialize the configuration manager.
        
        Args:
            config_path: Path to the configuration file
        """
        self.config_path = config_path
        self.config = None
        self._load_config()
    
    def _load_config(self) -> None:
        """Load configuration from file."""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
                self._validate_config()
            else:
                self.config = self._get_default_config()
                self.save_config()
        except Exception as e:
            print(f"Error loading config: {e}")
            print("Using default configuration...")
            self.config = self._get_default_config()
    
    def _get_default_config(self) -> Dict[str, Any]:
        """Get default configuration."""
        return {
            "organization": {
                "name": "Your Organization",
                "project_name": "Team Task Planner"
            },
            "teams": [
                {
                    "name": "Frontend Team",
                    "description": "User interface and frontend development",
                    "sample_tasks": []
                },
                {
                    "name": "Backend Team", 
                    "description": "Server-side development and APIs",
                    "sample_tasks": []
                },
                {
                    "name": "DevOps Team",
                    "description": "Infrastructure and deployment", 
                    "sample_tasks": []
                },
                {
                    "name": "QA Team",
                    "description": "Quality assurance and testing",
                    "sample_tasks": []
                }
            ],
            "task_headers": [
                "Task ID", "Task Name", "Assigned To", "Priority", "Status",
                "Start Date", "Due Date", "Progress %", "Hours Est.", "Hours Actual",
                "Dependencies", "Subtasks", "Notes"
            ],
            "status_options": ["Not Started", "Planning", "In Progress", "Completed", "On Hold"],
            "priority_options": ["Low", "Medium", "High", "Critical"]
        }
    
    def _validate_config(self) -> None:
        """Validate the loaded configuration."""
        required_keys = ["organization", "teams", "task_headers", "status_options", "priority_options"]
        for key in required_keys:
            if key not in self.config:
                raise ValueError(f"Missing required configuration key: {key}")
        
        # Validate organization
        org_keys = ["name", "project_name"]
        for key in org_keys:
            if key not in self.config["organization"]:
                raise ValueError(f"Missing organization key: {key}")
        
        # Validate teams
        if not isinstance(self.config["teams"], list) or len(self.config["teams"]) == 0:
            raise ValueError("Teams must be a non-empty list")
        
        for team in self.config["teams"]:
            if not isinstance(team, dict) or "name" not in team:
                raise ValueError("Each team must be a dictionary with a 'name' key")
    
    def save_config(self) -> None:
        """Save current configuration to file."""
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving config: {e}")
    
    def get_organization_name(self) -> str:
        """Get organization name."""
        return self.config["organization"]["name"]
    
    def get_project_name(self) -> str:
        """Get project name."""
        return self.config["organization"]["project_name"]
    
    def get_teams(self) -> List[Dict[str, Any]]:
        """Get list of teams."""
        return self.config["teams"]
    
    def get_team_names(self) -> List[str]:
        """Get list of team names."""
        return [team["name"] for team in self.config["teams"]]
    
    def get_team_by_name(self, name: str) -> Optional[Dict[str, Any]]:
        """Get team configuration by name."""
        for team in self.config["teams"]:
            if team["name"] == name:
                return team
        return None
    
    def get_task_headers(self) -> List[str]:
        """Get task headers."""
        return self.config["task_headers"]
    
    def get_status_options(self) -> List[str]:
        """Get status options."""
        return self.config["status_options"]
    
    def get_priority_options(self) -> List[str]:
        """Get priority options."""
        return self.config["priority_options"]
    
    def add_team(self, name: str, description: str = "", sample_tasks: List[Dict] = None) -> bool:
        """Add a new team to configuration.
        
        Args:
            name: Team name
            description: Team description
            sample_tasks: List of sample tasks
        
        Returns:
            True if team was added, False if team already exists
        """
        if self.get_team_by_name(name):
            return False
        
        new_team = {
            "name": name,
            "description": description,
            "sample_tasks": sample_tasks or []
        }
        
        self.config["teams"].append(new_team)
        self.save_config()
        return True
    
    def remove_team(self, name: str) -> bool:
        """Remove a team from configuration.
        
        Args:
            name: Team name to remove
        
        Returns:
            True if team was removed, False if team not found
        """
        for i, team in enumerate(self.config["teams"]):
            if team["name"] == name:
                del self.config["teams"][i]
                self.save_config()
                return True
        return False
    
    def update_organization(self, name: str = None, project_name: str = None) -> None:
        """Update organization information.
        
        Args:
            name: Organization name
            project_name: Project name
        """
        if name is not None:
            self.config["organization"]["name"] = name
        if project_name is not None:
            self.config["organization"]["project_name"] = project_name
        self.save_config()

def detect_existing_configuration(workbook_path: str) -> Optional[Dict[str, Any]]:
    """Detect existing configuration from an Excel workbook.
    
    Args:
        workbook_path: Path to existing Excel workbook
    
    Returns:
        Detected configuration or None if detection fails
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(workbook_path)
        
        # Extract organization name from Master Dashboard
        org_name = "Your Organization"
        if "Master Dashboard" in wb.sheetnames:
            master_sheet = wb["Master Dashboard"]
            title_cell = master_sheet['A1'].value
            if title_cell and " - " in str(title_cell):
                org_name = str(title_cell).split(" - ")[0].strip()
        
        # Extract team names (skip non-team sheets)
        skip_sheets = {"Master Dashboard", "Team Allocation", "Weekly Planning", "Templates"}
        teams = []
        for sheet_name in wb.sheetnames:
            if sheet_name not in skip_sheets:
                teams.append({
                    "name": sheet_name,
                    "description": f"{sheet_name} tasks and management",
                    "sample_tasks": []
                })
        
        config = {
            "organization": {
                "name": org_name,
                "project_name": "Team Task Planner"
            },
            "teams": teams,
            "task_headers": [
                "Task ID", "Task Name", "Assigned To", "Priority", "Status",
                "Start Date", "Due Date", "Progress %", "Hours Est.", "Hours Actual",
                "Dependencies", "Subtasks", "Notes"
            ],
            "status_options": ["Not Started", "Planning", "In Progress", "Completed", "On Hold"],
            "priority_options": ["Low", "Medium", "High", "Critical"]
        }
        
        return config
        
    except Exception as e:
        print(f"Error detecting configuration from workbook: {e}")
        return None

if __name__ == "__main__":
    # Test configuration manager
    config_manager = ConfigManager()
    print(f"Organization: {config_manager.get_organization_name()}")
    print(f"Teams: {config_manager.get_team_names()}")