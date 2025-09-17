#!/usr/bin/env python3
"""
Team Task Planner Excel Generator
Creates a comprehensive Excel workbook for managing multiple technical teams,
projects, deadlines, and task allocations.

Now supports dynamic configuration through config.json for flexible team setup.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import pandas as pd
from config_manager import ConfigManager

def create_team_task_planner(config_manager: ConfigManager = None):
    """Create the main Excel workbook with all required sheets.
    
    Args:
        config_manager: Configuration manager instance. If None, creates a new one.
    """
    
    if config_manager is None:
        config_manager = ConfigManager()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 1. Create Master Dashboard
    master_sheet = wb.create_sheet("Master Dashboard")
    create_master_dashboard(master_sheet, header_font, header_fill, border, center_align, config_manager)
    
    # 2. Create Team Allocation Sheet
    allocation_sheet = wb.create_sheet("Team Allocation")
    create_team_allocation(allocation_sheet, header_font, header_fill, border, center_align, config_manager)
    
    # 3. Create team sheets based on configuration
    teams = config_manager.get_teams()
    for team in teams:
        team_sheet = wb.create_sheet(team["name"])
        create_team_sheet(team_sheet, team, header_font, header_fill, subheader_font, 
                         subheader_fill, border, center_align, config_manager)
    
    # 4. Create Weekly Planning Sheet
    weekly_sheet = wb.create_sheet("Weekly Planning")
    create_weekly_planning(weekly_sheet, header_font, header_fill, border, center_align, config_manager)
    
    # 5. Create Templates Sheet
    templates_sheet = wb.create_sheet("Templates")
    create_templates_sheet(templates_sheet, header_font, header_fill, border, center_align, config_manager)
    
    # Set Master Dashboard as active sheet
    wb.active = master_sheet
    
    return wb

def create_master_dashboard(ws, header_font, header_fill, border, center_align, config_manager):
    """Create the master dashboard for overall project tracking."""
    
    ws.title = "Master Dashboard"
    
    # Title with organization name
    org_name = config_manager.get_organization_name()
    ws.merge_cells('A1:H2')
    ws['A1'] = f"{org_name.upper()} - TEAM TASK PLANNER - MASTER DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Current Date
    ws['A4'] = "Last Updated:"
    ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws['A4'].font = Font(bold=True)
    
    # Project Overview Headers
    headers = ["Project/Team", "Total Tasks", "Completed", "In Progress", "Pending", 
               "Overdue", "Team Size", "Workload %"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Dynamic team data based on configuration
    teams = config_manager.get_teams()
    teams_data = []
    
    for team in teams:
        # Create placeholder data for each team - this will be updated by formulas
        teams_data.append([
            team["name"],
            0,  # Total Tasks - will be calculated via formula
            0,  # Completed - will be calculated via formula
            0,  # In Progress - will be calculated via formula
            0,  # Pending - will be calculated via formula
            0,  # Overdue - will be calculated via formula
            0,  # Team Size - placeholder
            0   # Workload % - placeholder
        ])
    
    for row_idx, team_data in enumerate(teams_data, 7):
        for col_idx, value in enumerate(team_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = center_align
            
            # Color code workload percentage
            if col_idx == 8 and isinstance(value, (int, float)) and value > 0:  # Workload % column
                if value >= 90:
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif value >= 80:
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
    
    # Key Metrics with dynamic team count
    ws['A13'] = "KEY METRICS"
    ws['A13'].font = Font(size=14, bold=True)
    
    team_count = len(teams)
    end_row = 6 + team_count
    
    metrics = [
        ["Total Projects:", f"=COUNTA(A7:A{end_row})"],
        ["Total Tasks:", f"=SUM(B7:B{end_row})"],
        ["Overall Progress:", f"=IF(SUM(B7:B{end_row})>0,SUM(C7:C{end_row})/SUM(B7:B{end_row}),0)"],
        ["Critical Issues:", f"=SUM(F7:F{end_row})"],
        ["Average Team Load:", f"=IF(COUNT(H7:H{end_row})>0,AVERAGE(H7:H{end_row}),0)"]
    ]
    
    for row_idx, (metric, formula) in enumerate(metrics, 15):
        ws.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=formula)
        if "Progress" in metric:
            cell.number_format = "0.0%"
        elif "Load" in metric:
            cell.number_format = "0.0%"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_team_allocation(ws, header_font, header_fill, border, center_align, config_manager):
    """Create team allocation sheet for resource management."""
    
    ws.title = "Team Allocation"
    
    # Title
    ws.merge_cells('A1:I2')
    ws['A1'] = "TEAM RESOURCE ALLOCATION"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Employee Name", "Team", "Role", "Current Tasks", "Hours/Week", 
               "Availability %", "Skills", "Next Available", "Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Sample employee data - use first team as default for sample data
    teams = config_manager.get_teams()
    default_team = teams[0]["name"] if teams else "Default Team"
    
    employees = [
        ["John Doe", default_team, "Senior Developer", 3, 40, 90, "Programming, Analysis", "2024-01-15", "Busy"],
        ["Jane Smith", default_team, "Developer", 2, 35, 75, "Design, Development", "2024-01-10", "Available"],
    ]
    
    # Add sample employees for additional teams if they exist
    team_names = config_manager.get_team_names()
    if len(team_names) > 1:
        for i, team_name in enumerate(team_names[1:], 1):
            employees.extend([
                [f"Team Member {i*2-1}", team_name, "Developer", 2, 40, 70, "Various Skills", "2024-01-08", "Available"],
                [f"Team Member {i*2}", team_name, "Senior Developer", 3, 40, 85, "Leadership, Technical", "2024-01-12", "Busy"]
            ])
    
    for row_idx, emp_data in enumerate(employees, 5):
        for col_idx, value in enumerate(emp_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = center_align
            
            # Color code availability
            if col_idx == 6:  # Availability % column
                if value >= 90:
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif value >= 80:
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            
            # Color code status
            if col_idx == 9:  # Status column
                if value == "Available":
                    cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                elif value == "Busy":
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                elif value == "Overloaded":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    # Summary section
    ws['A13'] = "TEAM SUMMARY"
    ws['A13'].font = Font(size=14, bold=True)
    
    summary_headers = ["Team", "Total Members", "Available", "Busy", "Overloaded", "Avg Availability"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=15, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_team_sheet(ws, team, header_font, header_fill, subheader_font, 
                     subheader_fill, border, center_align, config_manager):
    """Create individual team task management sheet.
    
    Args:
        ws: Worksheet object
        team: Team configuration dictionary from config
        header_font, header_fill, subheader_font, subheader_fill, border, center_align: Styling
        config_manager: Configuration manager instance
    """
    
    team_name = team["name"]
    
    # Title
    ws.merge_cells('A1:M2')
    ws['A1'] = f"{team_name.upper()} - TASK MANAGEMENT"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Current Week
    ws['A4'] = f"Week of: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A4'].font = Font(bold=True)
    
    # Task Headers - use dynamic headers from configuration
    headers = config_manager.get_task_headers()
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Use sample tasks from configuration if available, otherwise create minimal samples
    tasks = []
    if "sample_tasks" in team and team["sample_tasks"]:
        for task in team["sample_tasks"]:
            task_row = [
                task.get("task_id", ""),
                task.get("task_name", ""),
                task.get("assigned_to", ""),
                task.get("priority", "Medium"),
                task.get("status", "Not Started"),
                task.get("start_date", ""),
                task.get("due_date", ""),
                task.get("progress", 0),
                task.get("hours_est", 0),
                task.get("hours_actual", 0),
                task.get("dependencies", ""),
                task.get("subtasks", ""),
                task.get("notes", "")
            ]
            tasks.append(task_row)
    else:
        # Create minimal sample task for new teams
        tasks = [
            [f"{team_name[:3].upper()}001", f"Sample Task for {team_name}", "Team Member", "Medium", "Not Started", 
             datetime.now().strftime("%Y-%m-%d"), (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), 
             0, 10, 0, "", "", "Add your tasks here"]
        ]
    
    for row_idx, task_data in enumerate(tasks, 7):
        for col_idx, value in enumerate(task_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Color code based on status and priority
            if col_idx == 4:  # Priority column
                if value == "High":
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif value == "Medium":
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            
            if col_idx == 5:  # Status column
                if value == "Completed":
                    cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                elif value == "In Progress":
                    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                elif value == "Planning":
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            if col_idx == 8:  # Progress % column
                cell.number_format = "0%"
                cell.value = value / 100  # Convert to decimal for percentage format
    
    # Team Summary
    ws['A12'] = "TEAM SUMMARY"
    ws['A12'].font = Font(size=14, bold=True)
    
    summary_data = [
        ["Total Tasks:", "=COUNTA(A7:A9)"],
        ["Completed:", "=COUNTIF(E7:E9,\"Completed\")"],
        ["In Progress:", "=COUNTIF(E7:E9,\"In Progress\")"],
        ["Average Progress:", "=AVERAGE(H7:H9)"]
    ]
    
    for row_idx, (label, formula) in enumerate(summary_data, 14):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=formula)
        if "Progress" in label:
            cell.number_format = "0%"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_weekly_planning(ws, header_font, header_fill, border, center_align, config_manager):
    """Create weekly planning sheet."""
    
    # Title
    ws.merge_cells('A1:H2')
    ws['A1'] = "WEEKLY PLANNING"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Week selector
    current_date = datetime.now()
    ws['A4'] = "Week Starting:"
    ws['B4'] = current_date.strftime("%Y-%m-%d")
    ws['A4'].font = Font(bold=True)
    
    # Days of the week
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    for col, day in enumerate(days, 2):
        cell = ws.cell(row=6, column=col)
        cell.value = day
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Team rows from configuration
    teams = config_manager.get_team_names()
    
    for row_idx, team in enumerate(teams, 7):
        cell = ws.cell(row=row_idx, column=1)
        cell.value = team
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.border = border
        
        # Add cells for each day
        for col in range(2, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            # Add some sample tasks
            if col == 2:  # Monday
                cell.value = "Sprint Planning"
            elif col == 6:  # Friday
                cell.value = "Sprint Review"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_templates_sheet(ws, header_font, header_fill, border, center_align, config_manager):
    """Create templates and instructions sheet."""
    
    # Title
    ws.merge_cells('A1:F2')
    ws['A1'] = "TEMPLATES & INSTRUCTIONS"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_align
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Instructions
    instructions = [
        "",
        "HOW TO USE THIS TEAM TASK PLANNER:",
        "",
        "1. MASTER DASHBOARD:",
        "   - Overview of all teams and projects",
        "   - Key metrics and performance indicators",
        "   - Automatic calculations from team sheets",
        "",
        "2. TEAM ALLOCATION:",
        "   - Resource management and availability",
        "   - Skills tracking and team member status",
        "   - Color-coded availability indicators",
        "",
        "3. TEAM SHEETS (Frontend, Backend, DevOps, QA):",
        "   - Individual task management for each team",
        "   - Progress tracking and status updates",
        "   - Dependencies and subtask management",
        "",
        "4. WEEKLY PLANNING:",
        "   - Week-by-week task scheduling",
        "   - Cross-team coordination",
        "   - Sprint planning support",
        "",
        "COLOR CODING:",
        "- Green: Available/Completed/Low Priority",
        "- Yellow: Busy/In Progress/Medium Priority", 
        "- Red: Overloaded/High Priority/Overdue",
        "- Blue: In Progress tasks",
        "- Gray: Not Started tasks",
        "",
        "FORMULAS USED:",
        "- Task counts: COUNTA functions",
        "- Progress calculations: AVERAGE and SUM",
        "- Status counting: COUNTIF functions",
        "",
        "TO ADD NEW TASKS:",
        "1. Go to appropriate team sheet",
        "2. Add new row with task details",
        "3. Use consistent formatting",
        "4. Update dependencies as needed",
        "",
        "TO ADD NEW TEAM MEMBERS:",
        "1. Go to Team Allocation sheet",
        "2. Add new row with member details",
        "3. Update team summary if needed",
        "",
        "BEST PRACTICES:",
        "- Update progress weekly",
        "- Review master dashboard regularly",
        "- Use consistent task naming",
        "- Track dependencies carefully",
        "- Update availability status",
    ]
    
    for row_idx, instruction in enumerate(instructions, 4):
        cell = ws.cell(row=row_idx, column=1)
        cell.value = instruction
        if instruction.startswith(("HOW TO USE", "COLOR CODING:", "FORMULAS USED:", "BEST PRACTICES:")):
            cell.font = Font(bold=True, size=12)
        elif instruction.endswith(":"):
            cell.font = Font(bold=True)
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 60

def main(config_path="config.json"):
    """Main function to create and save the Excel workbook.
    
    Args:
        config_path: Path to configuration file
    """
    print("Creating Team Task Planner Excel workbook...")
    
    # Initialize configuration manager
    config_manager = ConfigManager(config_path)
    
    print(f"Organization: {config_manager.get_organization_name()}")
    print(f"Teams: {', '.join(config_manager.get_team_names())}")
    
    # Create the workbook
    wb = create_team_task_planner(config_manager)
    
    # Generate filename based on organization
    org_name = config_manager.get_organization_name().replace(' ', '_')
    filename = f"{org_name}_Task_Planner.xlsx"
    wb.save(filename)
    
    print(f"Excel workbook '{filename}' created successfully!")
    print(f"The workbook contains the following sheets:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    return filename

if __name__ == "__main__":
    import sys
    config_path = sys.argv[1] if len(sys.argv) > 1 else "config.json"
    main(config_path)