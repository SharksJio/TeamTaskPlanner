#!/usr/bin/env python3
"""
Quick Setup Script for Team Task Planner
Helps users customize the Team Task Planner for their specific organization.
Enhanced with dynamic configuration support.
"""

import sys
import os
from create_team_task_planner import main as create_planner
from config_manager import ConfigManager

def interactive_setup():
    """Interactive setup wizard for customizing the Team Task Planner."""
    
    print("🚀 Welcome to Team Task Planner Setup!")
    print("="*50)
    
    # Get organization name
    org_name = input("Enter your organization name (optional): ").strip()
    if not org_name:
        org_name = "Your Organization"
    
    # Get team names
    print("\n📋 Let's set up your teams:")
    print("Enter team names (press Enter with empty name to finish):")
    
    teams = []
    team_num = 1
    while True:
        team_name = input(f"Team {team_num}: ").strip()
        if not team_name:
            break
        teams.append({
            "name": team_name,
            "description": f"{team_name} tasks and management",
            "sample_tasks": []
        })
        team_num += 1
    
    if not teams:
        print("No teams specified. Using default teams:")
        teams = [
            {"name": "Frontend Team", "description": "User interface and frontend development", "sample_tasks": []},
            {"name": "Backend Team", "description": "Server-side development and APIs", "sample_tasks": []},
            {"name": "DevOps Team", "description": "Infrastructure and deployment", "sample_tasks": []},
            {"name": "QA Team", "description": "Quality assurance and testing", "sample_tasks": []}
        ]
        for team in teams:
            print(f"  - {team['name']}")
    
    # Get filename
    filename = input(f"\n💾 Excel filename (default: {org_name.replace(' ', '_')}_Task_Planner.xlsx): ").strip()
    if not filename:
        filename = f"{org_name.replace(' ', '_')}_Task_Planner.xlsx"
    
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # Configuration filename
    config_filename = filename.replace('.xlsx', '_config.json')
    
    # Confirm setup
    print(f"\n✅ Setup Summary:")
    print(f"Organization: {org_name}")
    print(f"Teams: {', '.join(teams)}")
    print(f"Filename: {filename}")
    
    confirm = input("\nProceed with setup? (y/N): ").strip().lower()
    if confirm != 'y':
        print("Setup cancelled.")
        return
    
    # Create the planner
    print(f"\n🔧 Creating {filename}...")
    try:
        # Create basic planner first
        wb = create_planner()
        
        # Remove default sheets and add custom teams
        for sheet_name in list(wb.sheetnames):
            if "Team" in sheet_name and sheet_name not in ["Team Allocation"]:
                if sheet_name != "Frontend Team":  # Keep one as template
                    wb.remove(wb[sheet_name])
        
        # Rename the template and add custom teams
        if teams and "Frontend Team" in wb.sheetnames:
            frontend_sheet = wb["Frontend Team"]
            if teams[0] != "Frontend Team":
                frontend_sheet.title = teams[0]
                # Update title in the sheet
                frontend_sheet['A1'] = f"{teams[0].upper()} - TASK MANAGEMENT"
        
        # Add remaining teams
        from create_team_task_planner import create_team_sheet
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True)
        subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        for team in teams[1:]:
            team_sheet = wb.create_sheet(team)
            create_team_sheet(team_sheet, team, header_font, header_fill, subheader_font, 
                             subheader_fill, border, center_align)
        
        # Update organization name in Master Dashboard
        master_sheet = wb["Master Dashboard"]
        master_sheet['A1'] = f"{org_name.upper()} - TEAM TASK PLANNER - MASTER DASHBOARD"
        
        # Save the customized workbook
        wb.save(filename)
        
        print(f"✅ Successfully created {filename}!")
        print(f"\n📖 Next steps:")
        print(f"1. Open {filename} in Microsoft Excel")
        print(f"2. Start with the Master Dashboard for an overview")
        print(f"3. Add team members in the Team Allocation sheet")
        print(f"4. Begin adding tasks in your team sheets")
        print(f"5. Use Weekly Planning for sprint coordination")
        
        # Create quick start guide
        guide_filename = f"{org_name.replace(' ', '_')}_Quick_Start.txt"
        with open(guide_filename, 'w') as f:
            f.write(f"QUICK START GUIDE - {org_name} Team Task Planner\n")
            f.write("="*60 + "\n\n")
            f.write(f"Excel File: {filename}\n")
            f.write(f"Created: {os.getcwd()}\n\n")
            f.write("YOUR TEAMS:\n")
            for i, team in enumerate(teams, 1):
                f.write(f"{i}. {team}\n")
            f.write(f"\nGETTING STARTED:\n")
            f.write(f"1. Open {filename} in Excel\n")
            f.write(f"2. Review Master Dashboard\n")
            f.write(f"3. Add team members to Team Allocation sheet\n")
            f.write(f"4. Start adding tasks to team sheets\n")
            f.write(f"5. Plan weekly activities\n\n")
            f.write(f"For detailed documentation, see DOCUMENTATION.md\n")
        
        print(f"\n📄 Quick start guide created: {guide_filename}")
        
    except Exception as e:
        print(f"❌ Error creating planner: {e}")
        sys.exit(1)

def main():
    """Main function."""
    if len(sys.argv) > 1:
        if sys.argv[1] in ['-h', '--help', 'help']:
            print("Team Task Planner Setup")
            print("Usage:")
            print("  python3 setup.py          # Interactive setup")
            print("  python3 setup.py --help   # Show this help")
            return
    
    interactive_setup()

if __name__ == "__main__":
    main()