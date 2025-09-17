#!/usr/bin/env python3
"""
Team Task Planner Utilities
Additional utilities for customizing and working with the Team Task Planner Excel file.
Enhanced with dynamic configuration support.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import sys
import os
from config_manager import ConfigManager, detect_existing_configuration

def add_new_team_sheet(workbook_path, team_name, config_path="config.json"):
    """Add a new team sheet to an existing workbook using dynamic configuration.
    
    Args:
        workbook_path: Path to existing Excel workbook
        team_name: Name of the new team to add
        config_path: Path to configuration file
    
    Returns:
        True if team was added successfully, False otherwise
    """
    try:
        # Load configuration manager
        config_manager = ConfigManager(config_path)
        
        # Check if configuration exists from existing workbook
        if not os.path.exists(config_path):
            print("No configuration file found. Attempting to detect from existing workbook...")
            detected_config = detect_existing_configuration(workbook_path)
            if detected_config:
                config_manager.config = detected_config
                config_manager.save_config()
                print(f"Created configuration file: {config_path}")
            else:
                print("Could not detect configuration. Using defaults.")
        
        wb = openpyxl.load_workbook(workbook_path)
        
        # Check if team already exists
        if team_name in wb.sheetnames:
            print(f"Team '{team_name}' already exists!")
            return False
        
        # Add team to configuration
        if not config_manager.add_team(team_name, f"{team_name} tasks and management"):
            print(f"Team '{team_name}' already exists in configuration!")
        
        # Find an existing team sheet to use as template
        team_names = config_manager.get_team_names()
        template_team = None
        for existing_team in team_names:
            if existing_team in wb.sheetnames and existing_team != team_name:
                template_team = existing_team
                break
        
        if not template_team:
            print("No existing team sheet found to use as template!")
            return False
        
        # Create new sheet based on existing team template
        template_sheet = wb[template_team]
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = team_name
        
        # Update the title in the new sheet
        new_sheet['A1'] = f"{team_name.upper()} - TASK MANAGEMENT"
        
        # Clear sample data (keep headers) - find the right range based on headers
        # Headers are in row 6, so clear from row 7 onwards
        last_row = template_sheet.max_row
        for row in range(7, last_row + 1):
            for col in range(1, 14):  # Clear columns A-M
                cell = new_sheet.cell(row=row, column=col)
                if cell.value and not str(cell.value).startswith("="):  # Keep formulas
                    cell.value = ""
        
        # Add a sample task for the new team
        sample_task = [
            f"{team_name[:3].upper()}001",
            f"Sample Task for {team_name}",
            "Team Member",
            "Medium",
            "Not Started",
            datetime.now().strftime("%Y-%m-%d"),
            (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"),
            0, 10, 0, "", "", "Add your tasks here"
        ]
        
        for col_idx, value in enumerate(sample_task, 1):
            new_sheet.cell(row=7, column=col_idx).value = value
        
        wb.save(workbook_path)
        print(f"Successfully added team '{team_name}' to {workbook_path}")
        
        # Update master dashboard
        update_master_dashboard(workbook_path, config_manager.get_team_names(), config_path)
        
        return True
        
    except Exception as e:
        print(f"Error adding team: {e}")
        return False

def update_master_dashboard(workbook_path, teams_list, config_path="config.json"):
    """Update master dashboard with new team information.
    
    Args:
        workbook_path: Path to Excel workbook
        teams_list: List of team names
        config_path: Path to configuration file
    """
    try:
        wb = openpyxl.load_workbook(workbook_path)
        master_sheet = wb["Master Dashboard"]
        
        # Clear existing team data (start from row 7 and clear based on team count)
        max_clear_row = 7 + len(teams_list) + 5  # Add buffer
        for row in range(7, max_clear_row):
            for col in range(1, 9):
                master_sheet.cell(row=row, column=col).value = ""
        
        # Add new teams
        for idx, team in enumerate(teams_list, 7):
            master_sheet.cell(row=idx, column=1).value = team
            # Add placeholder formulas that can be updated manually
            master_sheet.cell(row=idx, column=2).value = 0  # Total Tasks
            master_sheet.cell(row=idx, column=3).value = 0  # Completed
            master_sheet.cell(row=idx, column=4).value = 0  # In Progress
            master_sheet.cell(row=idx, column=5).value = 0  # Pending
            master_sheet.cell(row=idx, column=6).value = 0  # Overdue
            master_sheet.cell(row=idx, column=7).value = 0  # Team Size
            master_sheet.cell(row=idx, column=8).value = 0  # Workload %
        
        # Update key metrics formulas
        team_count = len(teams_list)
        end_row = 6 + team_count
        
        # Find the metrics section and update formulas
        for row in range(15, 25):  # Metrics usually start around row 15
            cell_a = master_sheet.cell(row=row, column=1)
            if cell_a.value:
                if "Total Projects:" in str(cell_a.value):
                    master_sheet.cell(row=row, column=2).value = f"=COUNTA(A7:A{end_row})"
                elif "Total Tasks:" in str(cell_a.value):
                    master_sheet.cell(row=row, column=2).value = f"=SUM(B7:B{end_row})"
                elif "Overall Progress:" in str(cell_a.value):
                    master_sheet.cell(row=row, column=2).value = f"=IF(SUM(B7:B{end_row})>0,SUM(C7:C{end_row})/SUM(B7:B{end_row}),0)"
                elif "Critical Issues:" in str(cell_a.value):
                    master_sheet.cell(row=row, column=2).value = f"=SUM(F7:F{end_row})"
                elif "Average Team Load:" in str(cell_a.value):
                    master_sheet.cell(row=row, column=2).value = f"=IF(COUNT(H7:H{end_row})>0,AVERAGE(H7:H{end_row}),0)"
        
        wb.save(workbook_path)
        print(f"Successfully updated master dashboard in {workbook_path}")
        return True
        
    except Exception as e:
        print(f"Error updating master dashboard: {e}")
        return False

def auto_configure_from_workbook(workbook_path, config_path="config.json"):
    """Automatically create configuration from existing workbook.
    
    Args:
        workbook_path: Path to existing Excel workbook
        config_path: Path where to save the configuration
    
    Returns:
        True if configuration was created successfully, False otherwise
    """
    try:
        detected_config = detect_existing_configuration(workbook_path)
        if detected_config:
            # Save the configuration
            config_manager = ConfigManager(config_path)
            config_manager.config = detected_config
            config_manager.save_config()
            
            print(f"✅ Successfully created configuration from {workbook_path}")
            print(f"📄 Configuration saved to: {config_path}")
            print(f"🏢 Organization: {detected_config['organization']['name']}")
            print(f"👥 Teams found: {', '.join([team['name'] for team in detected_config['teams']])}")
            
            return True
        else:
            print(f"❌ Could not detect configuration from {workbook_path}")
            return False
            
    except Exception as e:
        print(f"❌ Error auto-configuring from workbook: {e}")
        return False

def export_team_tasks_to_csv(workbook_path, team_name, output_file):
    try:
        wb = openpyxl.load_workbook(workbook_path)
        
        if team_name not in wb.sheetnames:
            print(f"Team '{team_name}' not found!")
            return False
        
        sheet = wb[team_name]
        
        # Read headers and data
        headers = []
        for col in range(1, 14):  # Columns A-M
            headers.append(sheet.cell(row=6, column=col).value)
        
        tasks = []
        row = 7
        while sheet.cell(row=row, column=1).value:  # While there's a Task ID
            task = []
            for col in range(1, 14):
                task.append(sheet.cell(row=row, column=col).value)
            tasks.append(task)
            row += 1
        
        # Write to CSV
        import csv
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            writer.writerows(tasks)
        
        print(f"Successfully exported {team_name} tasks to {output_file}")
        return True
        
    except Exception as e:
        print(f"Error exporting tasks: {e}")
        return False

def generate_weekly_report(workbook_path, output_file):
    """Generate a weekly progress report."""
    try:
        wb = openpyxl.load_workbook(workbook_path)
        
        report_data = []
        report_data.append("WEEKLY PROGRESS REPORT")
        report_data.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        report_data.append("="*50)
        
        # Get master dashboard data
        master_sheet = wb["Master Dashboard"]
        report_data.append("\nOVERALL SUMMARY:")
        
        for row in range(7, 11):  # Team data rows
            team = master_sheet.cell(row=row, column=1).value
            if team:
                total_tasks = master_sheet.cell(row=row, column=2).value or 0
                completed = master_sheet.cell(row=row, column=3).value or 0
                in_progress = master_sheet.cell(row=row, column=4).value or 0
                workload = master_sheet.cell(row=row, column=8).value or 0
                
                report_data.append(f"\n{team}:")
                report_data.append(f"  Total Tasks: {total_tasks}")
                report_data.append(f"  Completed: {completed}")
                report_data.append(f"  In Progress: {in_progress}")
                report_data.append(f"  Workload: {workload}%")
        
        # Get team allocation summary
        allocation_sheet = wb["Team Allocation"]
        report_data.append("\nTEAM AVAILABILITY:")
        
        row = 5
        while allocation_sheet.cell(row=row, column=1).value:
            name = allocation_sheet.cell(row=row, column=1).value
            team = allocation_sheet.cell(row=row, column=2).value
            availability = allocation_sheet.cell(row=row, column=6).value or 0
            status = allocation_sheet.cell(row=row, column=9).value
            
            report_data.append(f"  {name} ({team}): {availability}% - {status}")
            row += 1
        
        # Write report
        with open(output_file, 'w', encoding='utf-8') as f:
            for line in report_data:
                f.write(line + '\n')
        
        print(f"Weekly report generated: {output_file}")
        return True
        
    except Exception as e:
        print(f"Error generating report: {e}")
        return False

def validate_workbook(workbook_path):
    """Validate the workbook structure and data."""
    try:
        wb = openpyxl.load_workbook(workbook_path)
        
        required_sheets = ["Master Dashboard", "Team Allocation", "Weekly Planning", "Templates"]
        issues = []
        
        # Check required sheets
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                issues.append(f"Missing required sheet: {sheet_name}")
        
        # Check master dashboard
        if "Master Dashboard" in wb.sheetnames:
            master_sheet = wb["Master Dashboard"]
            if not master_sheet['A1'].value or "MASTER DASHBOARD" not in master_sheet['A1'].value:
                issues.append("Master Dashboard title is missing or incorrect")
        
        # Check team allocation
        if "Team Allocation" in wb.sheetnames:
            allocation_sheet = wb["Team Allocation"]
            expected_headers = ["Employee Name", "Team", "Role", "Current Tasks"]
            for col, header in enumerate(expected_headers, 1):
                cell_value = allocation_sheet.cell(row=4, column=col).value
                if cell_value != header:
                    issues.append(f"Team Allocation header mismatch: expected '{header}', got '{cell_value}'")
        
        if issues:
            print("Validation Issues Found:")
            for issue in issues:
                print(f"  - {issue}")
            return False
        else:
            print("Workbook validation passed!")
            return True
            
    except Exception as e:
        print(f"Error validating workbook: {e}")
        return False

def main():
    """Main utility function with command-line interface."""
    if len(sys.argv) < 2:
        print("Team Task Planner Utilities")
        print("Usage:")
        print("  python3 utilities.py add_team <workbook_path> <team_name> [config_path]")
        print("  python3 utilities.py auto_config <workbook_path> [config_path]")
        print("  python3 utilities.py export_csv <workbook_path> <team_name> <output_file>")
        print("  python3 utilities.py weekly_report <workbook_path> <output_file>")
        print("  python3 utilities.py validate <workbook_path>")
        print("")
        print("New Dynamic Features:")
        print("  auto_config - Create configuration from existing workbook")
        print("  add_team - Add team using dynamic configuration")
        return
    
    command = sys.argv[1]
    
    if command == "add_team" and len(sys.argv) >= 4:
        workbook_path = sys.argv[2]
        team_name = sys.argv[3]
        config_path = sys.argv[4] if len(sys.argv) > 4 else "config.json"
        add_new_team_sheet(workbook_path, team_name, config_path)
        
    elif command == "auto_config" and len(sys.argv) >= 3:
        workbook_path = sys.argv[2]
        config_path = sys.argv[3] if len(sys.argv) > 3 else "config.json"
        auto_configure_from_workbook(workbook_path, config_path)
        
    elif command == "export_csv" and len(sys.argv) == 5:
        workbook_path = sys.argv[2]
        team_name = sys.argv[3]
        output_file = sys.argv[4]
        export_team_tasks_to_csv(workbook_path, team_name, output_file)
        
    elif command == "weekly_report" and len(sys.argv) == 4:
        workbook_path = sys.argv[2]
        output_file = sys.argv[3]
        generate_weekly_report(workbook_path, output_file)
        
    elif command == "validate" and len(sys.argv) == 3:
        workbook_path = sys.argv[2]
        validate_workbook(workbook_path)
        
    else:
        print("Invalid command or arguments. Use without arguments to see usage.")

if __name__ == "__main__":
    main()